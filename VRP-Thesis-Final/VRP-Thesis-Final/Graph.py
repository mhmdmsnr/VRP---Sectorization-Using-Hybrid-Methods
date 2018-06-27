from Vertix import Vertix
from Tools import Tools
from Sector import Sector
from customStr import customStr
import numpy as np
import Vertix
import csv
import math
import subprocess
import os
import openpyxl
import copy
import threading
import random
import folium

class Graph:
    """This class represents the graph in the solution whcih contains the sectors, costs, and routs
    
    Attributes:
    sectors[]: a list of the sectors contained in the graph
    freeVertices[]: a list of the vertices that are not yet obtained by any of the sectors
    depot: a vertix that represents the depot
    costs[][][]: a 3D array which represents the cost pairs between each two vertices for each of the criteria
    costsWeight[]: a list of the costs weight (importance of each of the criteria)
    verticesNumber: number of vertices included in the graph
    sectorsNumber: number of sectors in the graph
    pheromone[][]: the pheromone for each of the vertices and each of the sectors 
    """

    def __init__(self, numberOfSectors = int, verticesFilePath = str, criteriaFilePath = str, criteriaWeightFilePath = str, sectorMinSize = int, sectorMaxSize = int):
        self.sectorMaxSize = sectorMaxSize
        self.sectors = np.empty(numberOfSectors, dtype = Sector)
        self.setsectorsNumber(numberOfSectors)
        
        self.readVerticesCSV(verticesFilePath)                          #self.depot, and self.freeVertices are defined inside the function
        self.readCostsFile(criteriaFilePath, criteriaWeightFilePath)    #self.costs, and costs weight are defined inside the function
        self.updateCombinedCosts()
        self.pheromone = np.empty((self.verticesNumber + 1, self.sectorsNumber), dtype = float)
        self.sectorMinSize = sectorMinSize
        self.totalCost = 0

    def getdepot(self):
        return self.depot

    def getsectorMinSize(self):
        return self.sectorMinSize

    def moveVertix(self, vertix = Vertix, target = Sector):
        source = self.sectors[vertix.getsectorId()]
        source.freeVertix(vertix)
        target.obtainVertix(vertix)
        return 0

    def addVertixToSector(self, target = Sector, vertix = Vertix):
        target.obtainVertix(vertix)
        index = np.argwhere(self.freeVertices == vertix)
        self.freeVertices = np.delete(self.freeVertices, index)
        #for i in range (len(self.freeVertices)):
        #    if self.freeVertices[i].getid() == vertix.getid():
        #        self.freeVertices = np.delete(self.freeVertices, [i])
        #        break
        return 0

    def setcostsWeight(self, costIndex = int, value = float):
        self.costsWeight[costIndex] = value
        return 0

    def getcostsWeight(self, costIndex = int):
        return self.costsWeight[costIndex]

    def setcost(self, costIndex = int, source = Vertix, target = Vertix):
        self.costs[costIndex][source.getid()][target.getid()]
        return 0

    def getfreeVertices(self):
        return self.freeVertices

    #def updateSectorCombinedCost(self, sector = Sector):   #canceled because the combined cost is obtained from the CPLEX automatically
    #    return 0

    def updateSectorCosts(self, sector = Sector):
        sector.updateCosts(self.costs, self.depot)
        #self.de
        return 0

    def updateCostsForAllSectors(self):
        for sector in self.sectors:
            sector.updateCombinedCost(self.combinedCosts, self.getdepot())
            sector.updateCosts(self.costs, self.depot)
        return 0

    def readVerticesCSV(self, csvFilePath):
        csvFile = open(csvFilePath)
        fileReader = csv.reader(csvFile)
        counter = 0
        tempArray = np.empty([0,2], dtype = float)
        for row in fileReader:
            temp = np.empty([2])
            for i in range(2):
                temp[i] = row[i]
            counter += 1
            tempArray = np.append(tempArray, [temp], axis = 0)
        #fileReader = csv.reader(csvFile)
        self.setverticesNumber(counter - 1)
        self.freeVertices = np.empty(self.verticesNumber, dtype = Vertix.Vertix)
        #row = list(fileReader)[0]
        #row = fileReader.next()
        self.depot = Vertix.Vertix(-1 ,tempArray[0][0], tempArray[0][1], 0)
        for i in range(self.verticesNumber):
            self.freeVertices[i] = Vertix.Vertix(-1, tempArray[i + 1][0], tempArray[i + 1][1], i + 1)
            #if flag == 0:
            #    self.vertices.x = row[0]
            #    self.vertices.y = row[1]
            #    flag = 1
            #    continue
            #v = Vertix(row[0],row[1])
            #self.vertices = np.append(self.vertices,v)
        csvFile.close()
        return 0

    def readCostsFile(self, costsXlsxFilePath, costsWeightXlsxFilePath):
        ### this part takes care of the costsWeight array
        wb = openpyxl.load_workbook(costsWeightXlsxFilePath)
        ws = wb.get_sheet_by_name("Weights")
        #ws.rows()
        self.costsWeight = np.empty(ws.max_row, dtype = float)
        for i in range(ws.max_row):
            self.costsWeight[i] = ws.cell(row = i + 1, column = 1).value

        
        ### this part takes care of the costs[] array
        wb = openpyxl.load_workbook(costsXlsxFilePath)
        sheets = wb._sheets
        l = len(sheets)
        #m = self.numberOfVertices
        self.costs = np.ndarray(shape=(l, self.verticesNumber + 1, self.verticesNumber + 1), dtype=float)
        for i in range(len(self.costsWeight)):
            for j in range(self.verticesNumber):
                for k in range(self.verticesNumber):
                    self.costs[i][j][k] = sheets[i].cell(row = j + 1, column = k + 1).value

        return 0

    def CPLEXPrepare(self, datFilePath = str, xlsxFilePath = str):
        for i in range(self.sectorsNumber):
            datFileName = datFilePath + '-%s.dat' %(i)
            xlsxFileName = xlsxFilePath + '-%s.xlsx' %(i)
            self.sectors[i].CPLEXprep(self.combinedCosts, datFileName, xlsxFileName, self.depot)
        return 0

    def CPLEXCall(self, modFileName = str, datFilePath = str):
        #for i in range(self.sectorsNumber):
        #    datFileName = datFilePath + i
        try:
            return subprocess.call(["oplrun", modFileName, datFilePath])
        except:
            print ('do nothing')
        else:
            print ('CPLEX ERROR, DATFile: %s' %(datFilePath))

    def CPLEXResultsRead(self, xlsxFilePath = str):
        for i in range(self.sectorsNumber):
            #datFileName = datFilePath + i
            xlsxFileName = xlsxFilePath + '-%s.xlsx' %(i)
            self.sectors[i].readCPLEXResults(xlsxFileName)
        return 0

    def calculateRoutes(self, modeFilePath = str, datFilePath = str, xlsxFilePath = str):
        print ('CPLEX for %s.dat, and %s.xlsx started' %(datFilePath, xlsxFilePath))
        self.CPLEXPrepare(datFilePath, xlsxFilePath)

        list_thread = []
        total = 0
        for i in range(0, self.sectorsNumber):
            #print ('--------------Sector %s--------------------' %(i))
            #print ('Sector %s: %s' %(i, self.sectors[i]))
            #xlsxFile = filesNames + '-%s.xlsx' %(i)
            datFile = datFilePath + '-%s.dat' %(i)
            t = threading.Thread(target = self.CPLEXCall, args=("tsp-bruno.mod", datFile))
            list_thread.append(t)
            t.start()
            #total += temp
        #    print('CPLEX %s started!' %(i))
        #print ('all CPLEX instances started!             %s' %(total))

        for t in list_thread:
            t.join()

        self.CPLEXResultsRead(xlsxFilePath)
        print ('CPLEX for %s.dat, and %s.xlsx ended' %(datFilePath, xlsxFilePath))
        return 0

    def mapWrite(self, mapFilePath = str, radius = 1):
        colours = np.empty([9],customStr)                       #array of 9 distinctive colors for plotting on a map
        colours[0] = customStr('FF0000')
        colours[1] = customStr('00FF00')
        colours[2] = customStr('0000FF')
        colours[3] = customStr('FFFF00')
        colours[4] = customStr('00FFFF')
        colours[5] = customStr('FF00FF')
        colours[6] = customStr('00CC99')
        colours[7] = customStr('CC0033')
        colours[8] = customStr('000000')

        map = folium.Map(location=[38.9637, 35.2433], zoom_start=6)        #Create a map instance

        for sector in self.sectors:                               #add the centroids and neighbourhoods to the map
            #S = i
            sector.updateCentroid()
            C = '#%s' %(colours[sector.getid()].value)
            x = sector.centroid.getx()
            y = sector.centroid.gety()
            r = sector.diameter * radius * 1000
            #map.circle_marker(location=[x, y], radius=r, line_color=str(C), fill_color=str(C))
 
        for sector in self.sectors:                               #add the centroids and neighbourhoods to the map
            #S = i
            C = '#%s' %(colours[sector.getid()].value)
            x = sector.centroid.getx()
            y = sector.centroid.gety()
            #r = self.evaluateSector(i)[0] * radius * 1000
            #map.circle_marker(location=[x, y], radius=r, line_color=str(C), fill_color=str(C))
            map.polygon_marker(location=[x, y], popup=('%s, %s' %(str(sector.getid()), radius)), fill_color=str(C), num_sides=4, radius=10)

        map.polygon_marker(location=[self.depot.getx(), self.depot.gety()], popup="Depot", fill_color="#FFFFFF", num_sides=3, radius=8)
        
        for sector in self.sectors:
            vertices = sector.getvertices()
            sectorId = sector.getid()
            for j in range(len(vertices)):
                S = '%s, %s' %(sector.getid(), vertices[j].getid()) 
                C = '#%s' %(colours[sectorId].value)
                x = vertices[j].getx()
                y = vertices[j].gety()
                map.polygon_marker(location = [x, y], popup = str(S), fill_color = str(C), num_sides = 3, radius = 8)                

        map.create_map(mapFilePath)                                     #write the map to an HTML file
        firefox = "C://Program Files (x86)//Mozilla Firefox//firefox.exe"
        mapDir = "file:///D:/MhmdMnsr/Dropbox/Aref%20Mansour/Code/VRP-Thesis-Final/VRP-Thesis-Final/" + str(mapFilePath)
        #subprocess.call([firefox, mapDir])
        return 0

    def setverticesNumber(self, value = int):
        self.verticesNumber = value
        return 0

    def getverticesNumber(self):
        return self.verticesNumber

    def setsectorsNumber(self, value = int):
        self.sectorsNumber = value
        return 0

    def getsectorsNumber(self):
        return self.sectorsNumber

    def initiateSectors(self, sizePercentage = 1.2):
        self.sectors[0] = Sector(0, self.sectorMaxSize, len(self.costs), 0)
        vertices = self.getfreeVertices()
        furthestVertixIndex = 0
        maxDistance = Tools.distanceCalculationMethod(self.depot, vertices[0])
        for j in range(1, len(vertices)):
            temp = Tools.distanceCalculationMethod(self.depot, vertices[j])
            if maxDistance < temp:
                furthestVertixIndex = j
                maxDistance = temp
        self.addVertixToSector(self.sectors[0], vertices[furthestVertixIndex])
        
        for i in range(1, self.getsectorsNumber()):
            self.sectors[i] = Sector(0, self.sectorMaxSize, len(self.costs), i)
            vertices = self.getfreeVertices()
            furthestVertixIndex = 0
            maxDistance = Tools.distanceCalculationMethod(self.depot, vertices[0])
            for k in range(i):
                maxDistance += Tools.distanceCalculationMethod(self.sectors[k].getcentroid(), vertices[0])
            for j in range(1, len(vertices)):
                temp = Tools.distanceCalculationMethod(self.depot, vertices[j])
                for k in range(i):
                    temp += Tools.distanceCalculationMethod(self.sectors[k].getcentroid(), vertices[j])
                if maxDistance < temp:
                    furthestVertixIndex = j
                    maxDistance = temp
            self.addVertixToSector(self.sectors[i], vertices[furthestVertixIndex])
        self.mapWrite('test.html', 1)
        a = 1

        #--- at this point all the sectors have been initiated and are ready to obtain the rest of the free vertices in the graph
        for i in range(len(self.getfreeVertices())):
            min = 9999999999            #initial huge value to make sure it is bigger than any distance calculated
            vertixId = -1
            sectorId = -1
            freeVertices = self.getfreeVertices()
            for j in range(self.getsectorsNumber()):
                if (self.sectors[j].getsize() >= self.sectors[j].getlimit()):                     #i'm here trying to break the loop in case the sector is full already
                    continue
                vertices = self.sectors[j].getvertices()
                for k in range(len(vertices)):
                    for l in range(len(freeVertices)):
                        temp = Tools.distanceCalculationMethod(vertices[k], freeVertices[l])
                        if temp < min:
                            min = temp
                            vertixId = l
                            sectorId = j
            if vertixId == -1:
                print("error!")
                return -1
            self.addVertixToSector(self.sectors[sectorId], freeVertices[vertixId])

            #--- at this point the vertices will all be added to the sectors
        return 0              

    def initiateOptimization(self):
        for source in self.sectors:
            for vertix in source.getvertices():
                total = 0
                temp = np.empty(self.sectorsNumber)
                counter = 0
                for target in self.sectors:
                    temp[counter] = Tools.distanceCalculationMethod(vertix, target.getcentroid())
                    total += temp[counter]
                    counter += 1
                for i in range(self.sectorsNumber):
                    self.pheromone[vertix.getid(), i] = temp[i] / total
        return 0

    def getcombinedDiameters(self):
        temp = 0
        for sector in self.sectors:
            temp += sector.getdiameter()
        return temp

    def defineNeighbourhood(self, sector = Sector, numberOfcandidatesPerSector = int, neighbourhoodBoundarySize = 1.2):
        if sector.getsize() <= 1:
            if sector.getid() != 0:
                temp = self.sectors[0].getvertices()
            for i in range(1, self.getsectorsNumber()):
                if sector.getid() == i:
                    continue
                temp = np.append(temp, [self.sectors[i].getvertices()])
            return temp

        neighbourhoodBoundary = sector.getdiameter() * neighbourhoodBoundarySize
        neighbourhood = np.empty((0))
        for i in self.sectors:
            if i == sector:
                continue
            if i.getsize() <= self.sectorMinSize:
                continue
            temp = i.getFurthestVertices(numberOfcandidatesPerSector)
            if not hasattr(temp, "__iter__"):
                if Tools.distanceCalculationMethod(sector.getcentroid(), temp) <= neighbourhoodBoundary:
                    neighbourhood = np.append(neighbourhood, [temp])
            else:
                for j in temp:
                    if Tools.distanceCalculationMethod(sector.getcentroid(), j) < neighbourhoodBoundary:
                        neighbourhood = np.append(neighbourhood, [j])
        return neighbourhood        

    def globalPheromoneUpdate(self, value = float):
        for row in self.pheromone:
            for cell in row:
                if cell <= 0.05:
                    cell = 0.02
                else:
                    cell += value
        return 0

    def optimizationIteration(self, numberOfVerticesToObtain = 2, neighbourhoodSize = 1.05, localPheromoneUpdateValue = 0.05, globalPheromoneUpdateValue = -0.05):
        movementCounter = 0
        flags = np.zeros(self.getverticesNumber() + 1)
        for sectorId in range(self.getsectorsNumber()):
            neighbourhood = self.defineNeighbourhood(self.sectors[sectorId], numberOfVerticesToObtain, neighbourhoodSize)
            temp = np.empty(len(neighbourhood))
            for i in range(len(neighbourhood)):
                temp[i] = random.uniform(0,1) * self.pheromone[neighbourhood[i].getid(), sectorId]
            max = 0
            maxIndex = -1
            #print ('temp: %s' %(temp))
            for counter in range(numberOfVerticesToObtain):
                #print (counter)
                #if counter <= 0:
                #    break
                if self.sectors[sectorId].getlimit() <= self.sectors[sectorId].getsize():
                    break
                for i in range(len(temp)):
                    a = self.getsectorMinSize()
                    #if self.sectors[neighbourhood[i].getsectorId()].getsize() <= self.getsectorMinSize():
                    #    continue
                    if flags[neighbourhood[i].getid()] == 1:
                        continue
                    if temp[i] > max:
                        maxIndex = i
                        max = temp[i]

                if maxIndex == -1:
                    continue
                flags[neighbourhood[maxIndex].getid()] = 1
                movementCounter += 1
                self.moveVertix(neighbourhood[maxIndex], self.sectors[sectorId])
                self.localPheromoneUpdate(neighbourhood[maxIndex].getid(), sectorId, localPheromoneUpdateValue)
                #self.pheromone[neighbourhood[maxIndex].getid(), sectorId] += localPheromoneUpdateValue
                
        self.globalPheromoneUpdate(globalPheromoneUpdateValue)
        if movementCounter == 0:
            print ('No Vertix Obtainments have been done (neighbourhoodSize: %s)' %(neighbourhoodSize))
        return movementCounter

    def localPheromoneUpdate(self, vertixId = int, sectorId = int, changeValue = float):
        self.pheromone[vertixId, sectorId] += changeValue
        return 0

    def updateTotalCost(self):
        temp = 0
        for sector in self.sectors:
            temp += sector.getrouteCombinedCost()
        self.settotalCost(temp)
        return temp

    def settotalCost(self, Value):
        self.totalCost = Value
        return 0

    def gettotalCost(self):
        return self.totalCost

    def updateCombinedCosts(self):
        self.combinedCosts = np.empty(shape = (self.getverticesNumber() + 1, self.getverticesNumber() + 1), dtype = float)
        #vertices = self.getfreeVertices()
        for i in range(self.getverticesNumber() + 1):
            for j in range(self.getverticesNumber() + 1):
                for k in range(len(self.costsWeight)):
                    if i == j:
                        self.combinedCosts[i][i] = 0
                        continue
                    self.combinedCosts[i][j] = self.costs[k][i][j] * self.getcostsWeight(int(k))
        return 0

    def getweightsNumber(self):
        return len(self.costsWeight)

    def getsectorMaxSize(self):
        return self.sectorMaxSize