import numpy as np
import Vertix
import csv
import math
import subprocess
import os
import openpyxl
import copy
import threading
import random
import folium
from Tools import Tools
from Vertix import Vertix

class Sector:
    """this represents the sectors in the graph, which contain vertices.
    
    Attributes:
        size : current number of vertices included in this sector
        limit : upper limit of size for sector
        centroid : a vertix instance that represents the centroid of the sector
        diameter: the distance between centroid and the furthest vertix in the sector
        vertices[]: a list of the vertices indluded in this sector
        route[] : a list that represents the route of which the vehicle is supposed to traverse the vertices
        routeCosts[] : the cost of the proposed route according to each of the criteria specified
        routeCombinedCost: the combined cost of the criteria for the proposed route with the weights taken into consideration    
    """
    def getvertices(self):
        return self.vertices

    def __init__(self, size, limit, numberOfCriteria, id):
        self.id = id
        self.vertices = np.empty(shape = (0), dtype = Vertix)
        self.size = size
        self.limit = limit
        self.centroid = Vertix(self.id, 0, 0, 0)
        self.diameter = 0.0
        self.route = np.empty(size + 1, dtype = int)
        self.routeCosts = np.empty(numberOfCriteria, dtype = float)
        self.routeCombinedCost = 0.0 

    def getid(self):
        return self.id

    def obtainVertix(self, vertix = Vertix):
        """Obtains a free vertix and adds it to the list of vertices in the sector """

        if self.size >= self.limit:
            print ("Error!: sector reached limit")
            return 1
        vertix.setsectorId(self.getid())
        self.vertices = np.append(self.vertices, vertix)
        self.size += 1
        self.route = np.array(self.size + 1, dtype = int)
        self.updateDiameter()
        print ('sector: %s obtained vertix No. %s' %(self.getid(), vertix.getid()))
        return 0

    def freeVertix(self, vertixToDelete = Vertix):
        """frees a the vertix specified and removes it from the list of vertices included in the sector"""
        index = np.argwhere(self.vertices == vertixToDelete)
        self.vertices = np.delete(self.vertices, index)
        #for vertix in self.vertices:
        #    if vertix == vertixToDelete:
        #        self.vertices = np.delete(self.vertices, [vertix])
        #        break
        self.size -= 1
        return 0

    def getnumberOfVertices(self):
        return len(self.vertices)

    def updateCentroid(self):
        """Updates the centroid and the diameter of the current sector"""
        tempX = 0
        tempY = 0
        for vertix in self.vertices:
            tempX += vertix.getx()
            tempY += vertix.gety()
        self.centroid.setx(tempX / self.getnumberOfVertices())
        self.centroid.sety(tempY / self.getnumberOfVertices())

        max = Tools.distanceCalculationMethod(self.centroid, self.vertices[0])
        for vertix in self.vertices:
            temp = Tools.distanceCalculationMethod(self.centroid, vertix)
            if max < temp:
                max = temp
        self.diameter = max
        return 0

    def getsize(self):
        return self.size

    def setsize(self, value):
        self.size = value
        return 0

    def getlimit(self):
        return self.limit

    def setlimit(self, value):
        self.limit = value
        return 0

    def getFurthestVertix(self):
        furthestVertix = self.vertices[0]
        max = Tools.distanceCalculationMethod(self.getcentroid(), self.vertices[0])
        for vertix in self.vertices:
            temp = Tools.distanceCalculationMethod(self.getcentroid(), vertix)
            if max < temp:
                furthestVertix = vertix
                max = temp
        return furthestVertix

    def getFurthestVertices(self, numberOfVertices):
        if numberOfVertices == 1:
            return self.getFurthestVertix()
        if numberOfVertices >= self.getsize():
            return self.getvertices()
        list = np.empty((0), dtype = Vertix)
        flags = np.zeros(len(self.getvertices()))
        #counter = numberOfVertices
        for i in range(numberOfVertices):
            max = 0
            maxVertix = Vertix(0,0,0,0)
            counter = 0
            for vertix in self.vertices:
                if flags[counter] == 1:
                    break
                temp = Tools.distanceCalculationMethod(self.getcentroid(), vertix)
                if temp > max:
                    max = temp
                    maxVertix = vertix
                    flags[counter] = 1
            list = np.append(list, [maxVertix])
        return list


    def updateDiameter(self):
        self.updateCentroid()
        return 0

    def getcentroid(self):
        return self.centroid

    def getdiameter(self):
        return self.diameter

    def getroute(self):
        return self.route

    def setroute(self, value):
        self.route = value
        return 0

    def setrouteCost(self, costIndex = int, value = float):
        self.routeCosts[costIndex] = value
        return 0

    def getrouteCost(self, costIndex = int):
        return self.routeCosts[costIndex]

    def setrouteCombinedCost(self, value = float):
        self.routeCombinedCost = value
        return 0

    def getrouteCombinedCost(self):
        return self.routeCombinedCost

    def writeDatFile(self, datFilePath = str, xlsxFilePath = str):
        n = self.size + 1
        #print ('n: %s' %(n))
        txt = '//this file is created autmatically!\n'
        txt += 'n = %s; \n' %(n)
        txt += 'SheetConnection sheet("%s"); \n' %(xlsxFilePath)
        txt += 'depot from SheetRead(sheet,"\'vertices\'!A1"); \n'
        txt += 'indices from SheetRead(sheet,"\'vertices\'!A1:A%s"); \n' %(n)
        txt += 'edges1 from SheetRead(sheet,"\'distances\'!A1:C%s"); \n' %(n * (n-1))
        txt += 'solX to SheetWrite(sheet,"results!$A$1:$C$5000"); \n'
        txt += 'solU to SheetWrite(sheet,"order!$A$1:B$5000"); \n'
        txt += 'totalDistance to SheetWrite(sheet,"objective!$A$1"); \n'

        filePath = datFilePath
        with open(filePath,'w') as datFile:
            datFile.write(txt)
        return 0

    def writeXlsxFile(self,  costArray, xlsxFilePath = str, depot = Vertix):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "vertices"
        ws.append([depot.getid(), depot.getx(), depot.gety()])  #write the depot information
   
        for vertix in self.vertices:
            ws.append([vertix.getid(), vertix.getx(), vertix.gety()])

        #for i in range(0, int(self.sectorSize[sectorIndex])):   #write the vertices sheet
        #    vertixIndex = self.sectors[sectorIndex,i]
        #    ws.append([vertixIndex, self.vertices[vertixIndex].getx(), self.vertices[vertixIndex].gety()])
        
        wb.create_sheet("distances") 
        ws = wb.get_sheet_by_name("distances")

        for sourceVertix in self.vertices:                  #write the distances sheet
            sourceId = sourceVertix.getid()

            ws.append([depot.getid(), sourceId, costArray[depot.getid()][sourceId]])        #write the distance from the depot to the vertices
            ws.append([sourceId, depot.getid(), costArray[sourceId][depot.getid()]])        #write the distance to the depot from the vertices
                 
            for targetVertix in self.vertices:
                if sourceVertix == targetVertix:            #skip the distance between a vertix and itself
                    continue
                targetId = targetVertix.getid()
                ws.append([sourceId, targetId, costArray[sourceId][targetId]])
        
        wb.create_sheet("results") 
        wb.create_sheet("order") 
        wb.create_sheet("objective") 
        wb.save(xlsxFilePath)
        return 0

    def CPLEXprep(self, costsArray, datFilePath = str, xlsxFilePath = str, depot = Vertix):
        self.writeDatFile(datFilePath, xlsxFilePath)
        self.writeXlsxFile(costsArray, xlsxFilePath, depot)
        return 0

    def readCPLEXResults(self, xlsxFilePath = str):     #""", depot = Vertix"""
        self.route = np.empty(self.getsize() + 1, dtype = int)
        sheet = openpyxl.load_workbook(xlsxFilePath)
        #--- fetch routing sequence
        ws = sheet.get_sheet_by_name("order")
        #counter = 0
        for row in ws:                       
            self.route[(round(row[1].value - 1))] = row[0].value    #read route results and write it in the route array of the sector
            #counter += 1
        #self.route[counter] = depot.getid()                  #add the depot to the end of the route

        #--- fetch combined cost
        ws = sheet.get_sheet_by_name("objective")
        self.routeCombinedCost = ws['A1'].value
        return 0

    def updateCosts(self, costsArray, depot = Vertix):
        for i in range(len(self.routeCosts)):
            temp = 0
            temp += costsArray[i][depot.getid()][self.route[0]]
            for j in range(len(self.route) - 1):
                temp += costsArray[i][self.route[j]][self.route[j + 1]]
            temp += costsArray[i][self.route[self.size]][depot.getid()]
            self.setrouteCost(i, temp)
        return 0

    def updateCombinedCost(self, CombinedCostsArray, depot = Vertix):
        temp = 0
        for i in range(len(self.route) - 1):
            temp += CombinedCostsArray[self.route[i]][self.route[i + 1]]
        temp += CombinedCostsArray[self.route[len(self.route) - 1]][depot.getid()]
        self.setrouteCombinedCost(temp)
        return 0
