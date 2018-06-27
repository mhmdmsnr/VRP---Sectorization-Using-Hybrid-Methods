from Graph import Graph
from Vertix import Vertix
from Tools import Tools
from Sector import Sector
import numpy as np
import Vertix
import csv
import math
import subprocess
import os
import openpyxl
import copy
import threading
import random
import folium

class Solution:
    """This class represents the whole solution that should be ready for execution
    it includes the graph and all the elements needed for a solution as well as the score according to the criteria specified by the user
    
    Attributes:
    graph: the graph the corresponds to this solution
    score: solution's score
    totalCost: solution's combined cost (according to criteria and priorities given by user)
    routeCosts[]: array of costs for each of the criteria specified"""

    def __init__(self, graph = Graph):
        self.graph = graph

    def setfatherId(self, value = int):
        self.fatherId = value
        return 0

    def getfatherId(self):
        return self.fatherId

    def setlocalTime(self, value = float):
        self.localTime = value
        return 0

    def getlocalTime(self):
        return self.localTime

    def setglobalTime(self, value = float):
        self.globalTime = value
        return 0

    def getglobalTime(self):
        return self.globalTime

    def writeToXlsx(self, filePath = str):
        wb = openpyxl.Workbook()
        ws = wb.active

        #---write vertices---#
        ws.title = "vertices"
        ws.append([self.graph.getdepot().getid(), self.graph.getdepot().getx(), self.graph.getdepot().gety()])
        for sector in self.graph.sectors:
            for vertix in sector.getvertices():
                ws.append([vertix.getid(), vertix.getx(), vertix.gety()])
        #for i in range(self.graph.getdepot(), self.graph.getverticesNumber()):
        #    ws.append([i, self.graph.vertices[i].getx(), self.graph.vertices[i].gety()])  #write the depot information
        
        #---Write Routes---#
        wb.create_sheet("Routes") 
        ws = wb.get_sheet_by_name("Routes")
        totalRow = self.graph.sectorMaxSize + 5
        ws.cell(row = totalRow, column = self.graph.getsectorsNumber() + 1).value = self.graph.gettotalCost()       #write total cost of all routes
        ws.cell(row = totalRow, column = self.graph.getsectorsNumber() + 2).value = 'Total weighted Costs'       #writes the titles of criterias

        for i in range(self.graph.getweightsNumber()):
            ws.cell(row = totalRow + i + 1, column = self.graph.getsectorsNumber() + 3).value = 'Criteria #%s' %(i)      #writes the titles of criterias

        for sector in self.graph.sectors:
            ws.cell(row = 1, column = sector.getid() + 1).value = sector.getid()
            counter = 2
            for stop in sector.getroute():
                ws.cell(row = counter, column = sector.getid() + 1).value = str(stop)
                counter += 1
            for criterionIndex in range(self.graph.getweightsNumber()):
                ws.cell(row = totalRow + 1 + criterionIndex, column = sector.getid() + 1).value = sector.getrouteCost(criterionIndex)
            ws.cell(row = totalRow, column = sector.getid() + 1).value = sector.getrouteCombinedCost()

        wb.save(filePath)
        print ('%s saved!' %(filePath))

        return 0

        #totals = np.zeros((self.graph.getweightsNumber()))
        #for i in range(self.graph.getsectorsNumber()):    
        #    temp = np.zeros((self.graph.getweightsNumber()))
        #    for j in range(self.graph.getsectorMaxSize()):
        #        for k in range(self.graph.getweightsNumber()):
        #            if self.graph.routes[i,j] == 0:
        #                temp[k] += self.graph.criterias[k,self.graph.routes[i,j],self.graph.depot]         
        #                break
        #            if j == self.graph.sectorMaxSize-1:
        #                temp[k] += self.graph.criterias[k,self.graph.routes[i,j],self.graph.depot]         
        #                break
        #            temp[k] += self.graph.criterias[k,self.graph.routes[i,j],self.graph.routes[i,j+1]]
        #    for k in range(self.graph.numberOfWeights):
        #        totals[k] += temp[k]
        #        ws.cell(row=totalRow+k+1, column=i+1).value = temp[k]       
        #for k in range(int(len(totals))):
        #    ws.cell(row=totalRow+k+1, column=self.graph.numberOfSectors+1).value = totals[k]       
        
        #for i in range(1, self.graph.numberOfSectors+1):
        #    ws.cell(row=1, column=i).value = "sector "+str(i-1)
        #    ws.cell(row=totalRow, column=i).value = self.graph.costs[i-1]       #write total cost of a route
        #    #ws.cell(row=totalRow, column=self.graph.numberOfSectors+1).value = self.graph.totalCost       #write total cost of all routes
        #    for j in range(2, int(self.graph.sectorSize[i-1])+3):
        #        #print ('j: %s' %(j)) 
        #        ws.cell(row=j, column=i).value = self.graph.routes[i-1,j-2]

        
        ##---Write Costs---#
        #wb.save(fileName)
        #print ('file saved!')

        #return 0

    def isUnique(self, solutionList):
        for targetSolution in reversed(solutionList):
            counter = self.graph.getsectorsNumber()
            flags = np.zeros((counter))
            for sector in self.graph.sectors:
                for targetSector in targetSolution.graph.sectors:
                    if flags[targetSector.getid()] == 1:
                        continue
                    if sector.getsize() == targetSector.getsize():
                        counter -= 1
                        flags[targetSector.getid()] = 1
                        break
            if counter != 0:
                continue
            counter = self.graph.getsectorsNumber()
            sectorFlags = np.zeros((counter))
            for sector in self.graph.sectors:
                for targetSector in targetSolution.graph.sectors:
                    if sectorFlags[targetSector.getid()] == 1:
                        continue
                    if sector.getsize() != targetSector.getsize():
                        continue
                    vertixFlags = np.zeros((sector.getsize()))
                    vertixCounter = sector.getsize()
                    for vertix in sector.getvertices():
                        loopCounter = 0
                        for targetVertix in targetSector.getvertices():
                            if vertixFlags[loopCounter] == 1:
                                loopCounter += 1
                                continue
                            if vertix.getid() == targetVertix.getid():
                                vertixCounter -= 1
                                vertixFlags[loopCounter] = 1
                                break
                            loopCounter += 1                     
                    if vertixCounter == 0:
                        sectorFlags[targetSector.getid()] = 1
                        counter -= 1
                        break
            if counter == 0:
                return False
        print ('Unique Solution!')
        return True

    def pathFinder(self, solutionsList, ownIndex):
        if self.getfatherId() == None:
            return '-1 \n#%s,\t%s,\t%s,\t%s,' %(ownIndex, str(self.graph.gettotalCost()), str(self.getlocalTime()), str(self.getglobalTime()))         
        return '%s \n#%s,\t%s,\t%s,\t%s,' %(solutionsList[self.getfatherId()].pathFinder(solutionsList, self.getfatherId()), ownIndex, str(self.graph.gettotalCost()), str(self.getlocalTime()), str(self.getglobalTime()))
